'''
@File    :   manager.py
@Date    :   2021/01/05
@Author  :   Lucas Cortés Gutiérrez.
@Version :   4.0
@Contact :   lucas.cortes.14@sansano.usm.cl"
@Desc    :   Módulo de administración del sistema de reconocimineto de personas "Let Me In"
'''

# importar los paquetes necesarios.
from passlib.hash import sha256_crypt
from imutils.video import VideoStream
from tkinter import scrolledtext
from contextlib import suppress
from tkinter import messagebox
from imutils import paths
import face_recognition
from tkinter import *
import pyshine as ps
import numpy as np
import stdiomask
import openpyxl
import imutils
import pickle
import time
import cv2
import sys
import os

# Esta variable define si se entrena el sistema automáticamente luego del ingreso de un usuario nuevo.
auto_train = False

# Entrega feedback a través de CLI.
verbose = False

# Factor de escalamiento de la ventana que muestra la imagen de la cámara.
ventana = 0.8

# Función encargada del entrenamiento del modelo.
def train():
    if verbose: print("ENTRENANDO MODELO")
    all_face_encodings = {}
    # Lista todas las imágenes del dataset.
    imagePaths = list(paths.list_images('data/dataset'))

    # Recorre todas las imágenes del dataset.
    for i in imagePaths:
        name = i[13:-4]
        # Codifica las imágenes del dataset.
        all_face_encodings[name] = face_recognition.face_encodings(face_recognition.load_image_file(i),known_face_locations=None, num_jitters=10, model='large')[0]

    with open('data/model.dat', 'wb') as model:
        # Guarda el modelo entrenado.
        pickle.dump(all_face_encodings, model)
    if verbose: print("ENTRENAMIENTO FINALIZADO")

# Función encargada de generar una lista de los usuarios registrados.
def show_user(txt):
    # Abre el archivo de datos de usuario.
    wb = openpyxl.load_workbook("data/info.xlsx")
    ws = wb.active
    # Recorre el archivo y lo imprime fila por fila.
    for i in range(1,ws.max_row):
        nombre = ws.cell(row = i, column = 1).value
        depto = ws.cell(row = i, column = 2).value
        mail = ws.cell(row = i, column = 3).value
        debt = ws.cell(row = i, column = 4).value
        txt.insert(INSERT,nombre + " - " + depto + " - " + str(mail) + " - " + str(debt)+"\n")
        if verbose: print(nombre, " - ", depto)

# Función encargada de eliminar a un usuario de la base de datos.
def delete_user(usuario, depto):
    encontrado = False
    # Intenta remover la imagen del usuario a eliminar.
    try:
        os.remove("data/dataset/"+usuario+"-"+depto+".jpg")
        if verbose: print("IMAGEN DE USUARIO ELIMINADA")   
    except:
        print('\007')
        messagebox.showwarning('Error','Imagen de usuario no encontrada')
        if verbose: print("IMAGEN DE USUARIO NO ENCONTRADA")

    # Abre el archivo de datos de usuario.
    wb = openpyxl.load_workbook("data/info.xlsx")
    ws = wb.active

    # Recorre el archivo fila por fila.
    for i in range(1,ws.max_row):
        usr_cell = ws.cell(row = i, column = 1).value
        dep_cell = ws.cell(row = i, column = 2).value
        # Elimina la fila que contiene los datos del usuario.
        if (usr_cell == usuario and dep_cell == depto):
            encontrado = True
            ws.delete_rows(i,1)
            wb.save('data/info.xlsx')  
            messagebox.showwarning('Eliminar usuario','Usuario Eliminado') 
            if verbose: print("USUARIO ELIMINADO DE LA BASE DE DATOS")
            if auto_train: train() 
    if (not encontrado):
        print('\007')
        messagebox.showwarning('Eliminar usuario','Usuario no encontrado en la base de datos') 
        if verbose: print("USUARIO NO ENCONTRADO EN LA BASE DE DATOS")

# Función encargada de agregar un usuario a la base de datos.
def add_user(usuario, depto, mail, debt):
    # Abre el archivo de datos de usuario.
    wb = openpyxl.load_workbook("data/info.xlsx")
    ws = wb.active
    encontrado = False
    act_data = True
    capturar = True
    actualizar = False

    # Recorre el archivo y lo imprime fila por fila.
    for i in range(1,ws.max_row):
        usr_cell = ws.cell(row = i, column = 1).value
        dep_cell = ws.cell(row = i, column = 2).value
        mail_cell = ws.cell(row = i, column = 3).value
        debt_cell = ws.cell(row = i, column = 4).value

        # En caso de que el usuario a registrar ya se encuentre en la base de datos.
        if (usr_cell == usuario and dep_cell == depto):
            encontrado = True
            if verbose: print("USUARIO YA SE ENCUENTRA REGISTRADO")
            actualizar_datos = messagebox.askyesno('Usuario ya registrado','Usuario ya registrado\n ¿Desea actualizar las deudas y correos del usuario?')
            actualizar = True if actualizar_datos else False
            actualizar_foto = messagebox.askyesno('Usuario ya registrado','Usuario ya registrado\n ¿Desea actualizar la imagen del usuario?')
            capturar = True if actualizar_foto else False

    # Capturar de imagen de usuario.        
    if capturar:
        messagebox.showwarning('Captura fotográfica','Presionar Espacio para capturar, "ESC" para salir.')
        video = True
        #vs = VideoStream(src=0).start()
        # Inicia feed de video.
        vs = VideoStream(0).start()
        cv2.startWindowThread()
        cv2.namedWindow("REGISTRAR USUARIO")

        # Pausa para el inicio correcto del sensor de imagen.
        time.sleep(0.5)
        
        while video:
            # Variable que determina si existe una cara detectada.
            cara = False
            frame = vs.read()
            # Reflejar la imagen horizontalmente (espejo).
            frame = cv2.flip(frame, 1)
            picture = frame.copy()
            # Escalar la imagen a 1/4 de la original
            small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)

            rgb_small_frame = small_frame[:, :, ::-1]
            face_locations = face_recognition.face_locations(rgb_small_frame)

            for (top, right, bottom, left) in face_locations:
                top *= 4
                right *= 4
                bottom *= 4
                left *= 4
                picture = picture[top:bottom, left:right]

                # Proporción utilizada para desplazar los cuadros de texto respecto al cuadro que encierra la cara.
                proporcion = 8
                borde = proporcion
                escala = proporcion/10
                salto = proporcion * 5
                ancho = int((right-left)/proporcion)
                pos = bottom+20
                blanco = (255,255,255)
                verde = (0,255,200)
                azul = (0,100,255)
                rojo = (255,0,100)
                aceptado = azul
                denegado = rojo

                # Dibujar cuadrado.
                cv2.rectangle(frame, (left, top), (right, bottom), (250, 100, 0), 2)

                # Dibujar etiqueta con nombre.
                with suppress(Exception):
                    ps.putBText(frame, str(usuario), text_offset_x=left+ancho, text_offset_y=pos, vspace=borde, hspace=borde, font_scale=escala, background_RGB=aceptado, text_RGB=blanco)

                # Dibujar etiqueta con departamento.
                with suppress(Exception):
                    ps.putBText(frame, "Depto. "+str(depto), text_offset_x=left+ancho, text_offset_y=pos+50, vspace=borde, hspace=borde, font_scale=escala, background_RGB=aceptado, text_RGB=blanco)

                if len(face_locations) == 1:
                    cara = True  

            frame_show = frame.copy()
            # Texto de cierre de programa.
            ps.putBText(frame_show,'"ESC" para salir',text_offset_x=50,text_offset_y=frame_show.shape[0]-50,vspace=10,hspace=10, font_scale=1.0,background_RGB=(228,225,222),text_RGB=(1,1,1))
            
            # Texto de captura de imagen.
            ps.putBText(frame_show,'"Espacio" para capturar',text_offset_x=frame_show.shape[1]-450,text_offset_y=frame_show.shape[0]-50,vspace=10,hspace=10, font_scale=1.0,background_RGB=(228,225,222),text_RGB=(1,1,1))
            frame_show = cv2.resize(frame_show, (0, 0), fx=ventana, fy=ventana)

            cv2.imshow('REGISTRAR USUARIO', frame_show)
            #cv2.startWindowThread()

            key = cv2.waitKey(1) & 0xFF

            # Presionar "c" para capturar imagen
            #if key == ord('c') and cara == True: 
            if (key == ord(' ') or key == ord('c')) and cara == True: 
                cv2.imwrite('data/dataset/'+usuario+"-"+depto+'.jpg',picture)
                if verbose: print("IMAGEN CAPTURADA")
                cv2.destroyAllWindows()
                #VideoStream(0).stop()
                vs.stop()
                video = False
                #messagebox.showwarning('Captura fotográfica','Imagen capturada')
                break

            # Presionar "q" para salir
            #if key == ord("q")
            if key == 27:
                if verbose: print("ENROLAMIENTO CANCELADO")
                cv2.destroyAllWindows()
                #VideoStream(0).stop()
                vs.stop()
                video = False
                act_data = False
                print('\007')
                messagebox.showwarning('Captura fotográfica','Enrolamiento Cancelado')
                break

    if ((not encontrado and act_data) or actualizar):
        if actualizar:
            for i in range(1,ws.max_row):
                usr_cell = ws.cell(row = i, column = 1).value
                dep_cell = ws.cell(row = i, column = 2).value
                if (usr_cell == usuario and dep_cell == depto):
                    encontrado = True
                    ws.delete_rows(i,1)
                    wb.save('data/info.xlsx')  
        ws.insert_rows(2)
        ws.cell(row=2, column=1).value = usuario
        ws.cell(row=2, column=2).value = depto
        now = time.time()
        tiempo = time.localtime(now)
        time_log = time.strftime("%Y/%m/%d, %H:%M:%S", tiempo)
        ws.cell(row=2, column=5).value = time_log
        try:
            ws.cell(row=2, column=3).value = int(mail)
            ws.cell(row=2, column=4).value = int(debt)
        except:
            ws.cell(row=2, column=3).value = 0
            ws.cell(row=2, column=4).value = 0
        wb.save('data/info.xlsx')
        messagebox.showwarning('Agregar usuario','Usuario Agregado')
        if verbose: print("USUARIO AGREGADO A LA BASE DE DATOS")
        if auto_train: train()

