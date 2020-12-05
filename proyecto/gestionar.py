import train as train
from passlib.hash import sha256_crypt
from imutils.video import VideoStream
import face_recognition
import numpy as np
import openpyxl
import stdiomask
import getpass
import imutils
import json
import time
import cv2
import sys
import os

auto_train = True

with open("data/passwords.json", "r") as json_file: 
    data = json.load(json_file) 
    users = data["users"]
    passwords = data["passwords"]

#print("User 0 '{}', has password '{}'".format(users[0], passwords[0]))

def mostrar_user():
    wb = openpyxl.load_workbook("data/info.xlsx")
    ws = wb.active
    for i in range(1,ws.max_row):
        nombre = ws.cell(row = i, column = 1).value
        depto = ws.cell(row = i, column = 2).value
        print(nombre, " - ", depto)
    #gestion()

def eliminar_user():
    mostrar_user()
    encontrado = False
    usuario = input("Ingrese el nombre del usuario a eliminar : ").upper()
    depto = input("Ingrese el número de departamento : ")
    try:
        os.remove("data/img/"+usuario+"-"+depto+".jpg")
        print("IMAGEN DE USUARIO ELIMINADA")   
    except:
        print("IMAGEN DE USUARIO NO ENCONTRADA")

    wb = openpyxl.load_workbook("data/info.xlsx")
    ws = wb.active
    for i in range(1,ws.max_row):
        usr_cell = ws.cell(row = i, column = 1).value
        dep_cell = ws.cell(row = i, column = 2).value
        if (usr_cell == usuario and dep_cell == depto):
            encontrado = True
            ws.delete_rows(i,1)
            wb.save('data/info.xlsx')   
            print("USUARIO ELIMINADO DE LA BASE DE DATOS")
            train.train() 
    if (not encontrado):
        print("USUARIO NO ENCONTRADO EN LA BASE DE DATOS")

def agregar_user():
    mostrar_user()
    usuario = input("Ingrese el nombre del usuario a enrolar : ").upper()
    depto = input("Ingrese numero de departamento : ")
    encontrado = False

    wb = openpyxl.load_workbook("data/info.xlsx")
    ws = wb.active

    act_data = True
    for i in range(1,ws.max_row):
        usr_cell = ws.cell(row = i, column = 1).value
        dep_cell = ws.cell(row = i, column = 2).value
        if (usr_cell == usuario and dep_cell == depto):
            encontrado = True
            print("USUARIO YA SE ENCUENTRA REGISTRADO")
            actualizar = input("¿Desea actualizar la imagen del usuario? (Y) Si. (N) No. : ")
            if (actualizar == "Y" or actualizar == "y"):
                act_data = False 
            if (actualizar == "N" or actualizar == "n"):
                gestion()

    print("(C) Capturar Imagen. (Q) Salir. : ")
    video = True
    #vs = VideoStream(src=0).start()
    vs = VideoStream(0).start()
    cv2.startWindowThread()
    cv2.namedWindow("REGISTRAR USUARIO")
    #time.sleep(2.0)
    
    while video:
        cara = False
        frame = vs.read()
        frame = cv2.flip(frame, 1)
        picture = frame.copy()
        small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)

        rgb_small_frame = small_frame[:, :, ::-1]
        face_locations = face_recognition.face_locations(rgb_small_frame)

        for (top, right, bottom, left) in face_locations:
            top *= 4
            right *= 4
            bottom *= 4
            left *= 4
            picture = picture[top:bottom, left:right]
            cv2.rectangle(frame, (left, top), (right, bottom), (255, 0, 0), 2)
            if len(face_locations) == 1:
                cara = True      
        frame_show = frame.copy()
        frame_show = cv2.resize(frame, (0, 0), fx=0.5, fy=0.5)
        cv2.imshow('REGISTRAR USUARIO', frame_show)
        #cv2.startWindowThread()

        key = cv2.waitKey(1) & 0xFF

        # Presionar "c" para capturar imagen
        if key == ord('c') and cara == True: 
            cv2.imwrite('data/img/'+usuario+"-"+depto+'.jpg',picture)
            print("IMAGEN CAPTURADA")
            train.train()
            cv2.destroyAllWindows()
            #VideoStream(0).stop()
            vs.stop()
            video = False
            break

        # Presionar "q" para salir
        if key == ord("q"):
            print("ENROLAMIENTO CANCELADO")
            cv2.destroyAllWindows()
            #VideoStream(0).stop()
            vs.stop()
            video = False
            gestion()

    if (not encontrado and act_data):
        ws.insert_rows(2)
        ws.cell(row=2, column=1).value = usuario
        ws.cell(row=2, column=2).value = depto
        ws.cell(row=2, column=3).value = 0
        ws.cell(row=2, column=4).value = 0
        wb.save('data/info.xlsx')
        print("USUARIO AGREGADO A LA BASE DE DATOS")
        gestion()


def gestion():
    while True:
        modo = input("(A) Agregar/Actualizar Usuario. (E) Eliminar Usuario. (M) Mostrar Usuarios. (Q) Salir. : ")
        if (modo == "Q" or modo == "q"):
            print ("HASTA PRONTO")
            exit()
        if (modo == "E" or modo == "e"):
            eliminar_user()
        if (modo == "A" or modo == "a"):
            agregar_user()
        if (modo == "M" or modo == "m"):
            mostrar_user()
        if (modo == "P" or modo == "p"):
            main()

def gestion_check():
    while True:
        user = input("INGRESE USUARIO: ")
        while user in users:
            password = stdiomask.getpass("INGRESE CONTRASEÑA: ")
            for u in range(len(users)):
                if users[u] == user:
                    if sha256_crypt.verify(password, passwords[u]):
                        print ("BIENVENIDO", user)
                        gestion()
                        #break
                    else:
                        print ("CONTRASEÑA INCORRECTA")
                        gestion_check()
        else:
            print("USUARIO NO REGISTRADO") 

def gestion_gui(usr_gui, psw_gui):
    gestion()
    for u in range(len(users)):
        if users[u] == usr_gui:
            if sha256_crypt.verify(psw_gui, passwords[u]):
                gestion()

#gestion_check()    