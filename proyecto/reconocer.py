from imutils import paths
import face_recognition
import numpy as np
import argparse
import openpyxl
import cv2
import os

#ap = argparse.ArgumentParser()
#ap.add_argument("-n", "--name", required=True,
#    help="Nombre de la persona a ingresar")
#ap.add_argument("-p", "--prototxt", required=True,
#    help="path to Caffe 'deploy' prototxt file")
#ap.add_argument("-m", "--model", required=True,
#    help="path to Caffe pre-trained model")
#ap.add_argument("-c", "--confidence", type=float, default=0.5,
#    help="minimum probability to filter weak detections")
#args = vars(ap.parse_args())

video_capture = cv2.VideoCapture(0)
imagePaths = list(paths.list_images('dataset'))
known_face_encodings = []
known_face_names = []

wb = openpyxl.load_workbook("info.xlsx")
ws = wb.active

'''
for i in imagePaths:
    name = i.strip("/dataset").strip(".jpg").strip("_")
    image = str(i)+"_image"
    vars()[image] = face_recognition.load_image_file(i)
    encoding = str(i)+"_face_encoding"
    vars()[encoding] = face_recognition.face_encodings(vars()[image])[0]
    known_face_encodings.append(vars()[encoding])
    known_face_names.append(name)
'''
for i in imagePaths:
    #name = i.strip("/dataset").strip(".jpg").strip("_")
    name = i[8:-4]
    known_face_encodings.append(face_recognition.face_encodings(face_recognition.load_image_file(i))[0])
    known_face_names.append(name)
    print(name)

# Initialize some variables
face_locations = []
face_encodings = []
face_names = []
process_this_frame = True


while True:
    # Grab a single frame of video
    ret, frame = video_capture.read()
    frame = cv2.flip(frame, 1)

    # Resize frame of video to 1/4 size for faster face recognition processing
    small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)

    # Convert the image from BGR color (which OpenCV uses) to RGB color (which face_recognition uses)
    rgb_small_frame = small_frame[:, :, ::-1]

    # Only process every other frame of video to save time
    if process_this_frame:
        # Find all the faces and face encodings in the current frame of video
        face_locations = face_recognition.face_locations(rgb_small_frame)
        face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

        face_names = []
        for face_encoding in face_encodings:
            # See if the face is a match for the known face(s)
            matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
            name = "Desconocido"
            

            # # If a match was found in known_face_encodings, just use the first one.
            # if True in matches:
            #     first_match_index = matches.index(True)
            #     name = known_face_names[first_match_index]

            # Or instead, use the known face with the smallest distance to the new face
            face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
            best_match_index = np.argmin(face_distances)
            if matches[best_match_index]:
                name = known_face_names[best_match_index]

            face_names.append(name)

    process_this_frame = not process_this_frame

    # Display the results
    for (top, right, bottom, left), name in zip(face_locations, face_names):
        depto = "Desconocido"
        correo = "0"
        deudas = "0"

        for j in range(1,ws.max_row):
            if (ws.cell(row = j, column = 1).value == name):
                depto = str(ws.cell(row = j, column = 2).value)
                correo = str(ws.cell(row = j, column = 3).value)
                deudas = str(ws.cell(row = j, column = 4).value)

        # Scale back up face locations since the frame we detected in was scaled to 1/4 size
        top *= 4
        right *= 4
        bottom *= 4
        left *= 4
        font = cv2.FONT_HERSHEY_DUPLEX
        if (name == "Desconocido"):
            # Draw a box around the face
            cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)

            # Draw a label with a name below the face
            cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 0, 255), cv2.FILLED)
            
        else:
            # Draw a box around the face
            cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)

            # Draw a label with a name below the face
            cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 255, 0), cv2.FILLED)

            if (depto == "Desconocido"):
                cv2.rectangle(frame, (left, bottom+35), (right, bottom), (0, 0, 255), cv2.FILLED)
            if (depto != "Desconocido"):
                cv2.rectangle(frame, (left, bottom+35), (right, bottom), (0, 255, 0), cv2.FILLED)

            cv2.putText(frame, "Depto. "+depto, (left + 6, bottom + 30), font, 1.0, (255, 255, 255), 1)

            if (int(correo) > 0):
                cv2.rectangle(frame, (left, bottom + 35), (right, bottom + 65), (0, 0, 255), cv2.FILLED)
                cv2.putText(frame, "Correo "+correo, (left + 6, bottom + 60), font, 1.0, (255, 255, 255), 1)

            if (int(deudas) > 0):
                cv2.rectangle(frame, (left, bottom + 65), (right, bottom + 95), (0, 0, 255), cv2.FILLED)
                cv2.putText(frame, "Deudas "+deudas, (left + 6, bottom + 90), font, 1.0, (255, 255, 255), 1)


        cv2.putText(frame, name, (left + 6, bottom - 6), font, 1.0, (255, 255, 255), 1)


    # Display the resulting image
    cv2.imshow('Video', frame)

    # Hit 'q' on the keyboard to quit!
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# Release handle to the webcam
video_capture.release()
cv2.destroyAllWindows()
