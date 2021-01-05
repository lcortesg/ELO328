'''
@File    :   monitor.py
@Date    :   2021/01/05
@Author  :   Lucas Cortés Gutiérrez.
@Version :   4.0
@Contact :   lucas.cortes.14@sansano.usm.cl"
@Desc    :   Módulo de monitoreo del sistema de reconocimineto de personas "Let Me In"
'''

# importar los paquetes necesarios
from contextlib import suppress
from tkinter import messagebox
from imutils import paths
import face_recognition
from tkinter import *
import pyshine as ps
import numpy as np
import openpyxl
import pickle
import time
import cv2
import os

tolerance = 0.5
tiempo_log  = 60
ventana = 0.8
name = ""
dpto = ""
correo = ""
deudas = ""

# Función encargada de la detención de la función "monitor()".
def monitor_kill():
    # Libera la cámara y destruye las ventanas
    cv2.VideoCapture(0).release()
    #video_capture.release()
    cv2.destroyAllWindows()

# Función encargada de logear los usuarios detectados.
def user_log(name, depto, frame, now):
    tiempo = time.localtime(now)
    time_log = time.strftime("%Y/%m/%d, %H:%M:%S", tiempo)
    time_pic = time.strftime("%Y-%m-%d, %H-%M-%S", tiempo)
    cv2.imwrite('data/log_user/'+time_pic+" - "+name+"-"+depto+'.jpg',frame)

    with open('data/log_user/log_user.txt', 'r') as original: data = original.read()

    if (name == "DESCONOCIDO" or name == "DUPLICADO"):
        with open('data/log_user/log_user.txt', 'w') as modified: modified.write(time_log+" - "+name+"\r\n" + data)

    else:    
        with open('data/log_user/log_user.txt', 'w') as modified: modified.write(time_log+" - "+name+"-"+depto+"\r\n" + data)

    original.close
    modified.close
    
# Función encargada del monitoreo de usuarios.
def monitor():
    ultimos = []
    before = time.time()

    with open('data/model.dat', 'rb') as f:
        all_face_encodings = pickle.load(f)
    
    video_capture = cv2.VideoCapture(0)
    
    wb = openpyxl.load_workbook("data/info.xlsx")
    ws = wb.active

    known_face_names = list(all_face_encodings.keys())
    known_face_encodings = np.array(list(all_face_encodings.values()))

    face_locations = []
    face_encodings = []
    face_names = []
    process_this_frame = True

    time.sleep(0.5)
    
    while True:
        ret, frame = video_capture.read()
        frame = cv2.flip(frame, 1)

        # Escalar la imagen a 1/4 de la original
        small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)

        # convertir la imagen desde BGR a RGB
        rgb_small_frame = small_frame[:, :, ::-1]

        # Se procesa 1 de cada 2 frames
        if process_this_frame:
            # Encuentra todas las caras en un frame dado
            face_locations = face_recognition.face_locations(rgb_small_frame)
            face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)
            face_names = []

            for face_encoding in face_encodings:
                # Revisar coincidencias
                matches = face_recognition.compare_faces(known_face_encodings, face_encoding, tolerance)
                name = "DESCONOCIDO"

                # # Se utiliza la primera coincidencia encontrada en known_face_encodings.
                # if True in matches:
                #     first_match_index = matches.index(True)
                #     name = known_face_names[first_match_index]

                # Se utiliza la mejor coincidencia a la cara detectada.
                face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
                best_match_index = np.argmin(face_distances)

                if matches[best_match_index]:
                    name = known_face_names[best_match_index]

                if name not in face_names:
                    face_names.append(name)

                if name in face_names:
                    face_names.append("DUPLICADO")

        process_this_frame = not process_this_frame

        # Muestra los resultados
        for (top, right, bottom, left), name in zip(face_locations, face_names):
            depto = "DESCONOCIDO"
            correo = 0
            deudas = 0
            user = name.split("-")
            
            for j in range(1,ws.max_row):
                if (ws.cell(row = j, column = 1).value == user[0] and ws.cell(row = j, column = 2).value == user[1]):
                    name = user[0]
                    depto = user[1]
                    correo = str(ws.cell(row = j, column = 3).value)
                    deudas = str(ws.cell(row = j, column = 4).value)

            # Desescalar la imagen
            top *= 4
            right *= 4
            bottom *= 4
            left *= 4
            font = cv2.FONT_HERSHEY_DUPLEX

            proporcion = 8
            borde = proporcion
            escala = proporcion/10
            salto = proporcion * 5
            ancho = int((right-left)/proporcion)
            pos = bottom+20
            blanco = (255,255,255)
            verde = (0,255,200)
            azul = (0,100,255)
            rojo = (255,0,100)
            naranjo = (255,100,0)
            aceptado = verde
            denegado = rojo

            if (name == "DUPLICADO"):
                # Dibujar cuadrado
                cv2.rectangle(frame, (left, top), (right, bottom), (0, 100, 255), 2)

                # Dibujar etiqueta con nombre
                with suppress(Exception):
                    ps.putBText(frame, name, text_offset_x=left+ancho, text_offset_y=bottom+20, vspace=borde, hspace=borde, font_scale=escala, background_RGB=naranjo, text_RGB=(255,250,250))

            elif (name == "DESCONOCIDO"):
                # Dibujar cuadrado
                cv2.rectangle(frame, (left, top), (right, bottom), (100, 0, 255), 2)

                # Dibujar etiqueta con nombre
                with suppress(Exception):
                    ps.putBText(frame, name, text_offset_x=left+ancho, text_offset_y=bottom+20, vspace=borde, hspace=borde, font_scale=escala, background_RGB=denegado, text_RGB=(255,250,250))
                
            else:
                # Dibujar cuadrado
                cv2.rectangle(frame, (left, top), (right, bottom), (200, 250, 0), 2)

                # Dibujar etiqueta con nombre
                with suppress(Exception):
                    ps.putBText(frame, str(name), text_offset_x=left+ancho, text_offset_y=pos, vspace=borde, hspace=borde, font_scale=escala, background_RGB=aceptado, text_RGB=blanco)

                # Dibujar etiqueta con departamento
                with suppress(Exception):
                    pos += salto
                    ps.putBText(frame, "Depto. "+str(depto), text_offset_x=left+ancho, text_offset_y=pos, vspace=borde, hspace=borde, font_scale=escala, background_RGB=aceptado, text_RGB=blanco)
                    
                if (int(correo) > 0):
                    # Dibujar etiqueta con correo
                    with suppress(Exception):
                        pos += salto
                        ps.putBText(frame, "Correo "+str(correo), text_offset_x=left+ancho, text_offset_y=pos, vspace=borde, hspace=borde, font_scale=escala, background_RGB=denegado, text_RGB=blanco)

                if (int(deudas) > 0):
                    # Dibujar etiqueta con deudas
                    with suppress(Exception):
                        pos += salto
                        ps.putBText(frame, "Deudas "+str(deudas), text_offset_x=left+ancho, text_offset_y=pos, vspace=borde, hspace=borde, font_scale=escala, background_RGB=denegado, text_RGB=blanco)
                            
        # Lógica de invocación de función "log_user()"
        now = time.time()
        with suppress(Exception):
            if (name+"-"+depto) not in ultimos:
                user_log(name, depto, frame, now)
                ultimos.append(name+"-"+depto)
            if int(now - before) >= tiempo_log:
                before = time.time()
                ultimos.clear()

        # Texto de cierre de programa.
        ps.putBText(frame,'"ESC" para salir',text_offset_x=50,text_offset_y=frame.shape[0]-50,vspace=10,hspace=10, font_scale=1.0,background_RGB=(228,225,222),text_RGB=(1,1,1))
        # Escala la imagen para mostrarla en pantalla en un tamaño más razonable
        frame = cv2.resize(frame, (0, 0), fx=ventana, fy=ventana)
        # Muestra imagen resultante
        cv2.imshow('Video', frame)

        # Presionar "q" para salir
        #if cv2.waitKey(1) & 0xFF == ord('q'): break
        if cv2.waitKey(1) & 0xFF == 27:
            res = messagebox.askokcancel('Salir','¿Detener monitoreo?')
            #res = messagebox.askyesno('Salir','¿Está seguro que desea detener el monitoreo')
            if res: break

    # Libera la cámara y destruye las ventanas
    video_capture.release()
    cv2.destroyAllWindows()
    
#monitor()